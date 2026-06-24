"""
Excel 导出服务
基于 openpyxl 生成 .xlsx 工作簿，支持商品列表、分析报告和利润报表

样式说明：
  - 表头：深蓝底白字加粗
  - 数据行：交替浅灰/白色
  - 数字列：右对齐，金额保留两位小数
  - 标题行：合并单元格 + 大字体
"""

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

from app.services.product_service import product_service

# ========== 样式定义 ==========
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1E3A5F")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="666666")
DATA_FONT = Font(name="微软雅黑", size=10)
NUMBER_FONT = Font(name="Consolas", size=10)
BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALT_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")


def _style_header(ws, row: int, cols: int):
    """为表头行应用样式"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN


def _style_data(ws, start_row: int, end_row: int, cols: int):
    """为数据区域应用样式"""
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = BORDER_THIN
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def _auto_width(ws, cols: int, min_width: int = 10, max_width: int = 50):
    """自动调整列宽"""
    for col in range(1, cols + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    # 中文字符宽度约等于 2
                    val_str = str(cell_val)
                    length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_len = max(max_len, length + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len, max_width)


class ExportService:
    """Excel 导出服务"""

    # ========== 1. 导出商品列表 ==========

    def export_products(self) -> io.BytesIO:
        """导出全部商品数据为 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "商品列表"

        products = product_service.get_all_raw_products()

        # 标题行
        ws.merge_cells("A1:J1")
        title_cell = ws.cell(row=1, column=1, value="AI 跨境电商智能选品分析平台 - 商品数据导出")
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:J2")
        sub_cell = ws.cell(row=2, column=1, value=f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}  |  商品总数：{len(products)}")
        sub_cell.font = SUBTITLE_FONT
        sub_cell.alignment = Alignment(horizontal="center")

        # 表头
        headers = ["ID", "商品标题", "类目", "价格($)", "销量", "评分", "评论数", "店铺名称", "来源平台", "描述"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=h)
        _style_header(ws, 4, len(headers))

        # 数据行
        for i, p in enumerate(products):
            row = 5 + i
            ws.cell(row=row, column=1, value=p["id"])
            ws.cell(row=row, column=2, value=p["title"])
            ws.cell(row=row, column=3, value=p["category"])
            ws.cell(row=row, column=4, value=p["price"])
            ws.cell(row=row, column=5, value=p["sales"])
            ws.cell(row=row, column=6, value=p["rating"])
            ws.cell(row=row, column=7, value=p["reviews"])
            ws.cell(row=row, column=8, value=p["store"])
            ws.cell(row=row, column=9, value=p.get("source_platform", ""))
            ws.cell(row=row, column=10, value=p.get("description", ""))

        end_row = 4 + len(products)
        _style_data(ws, 5, end_row, len(headers))

        # 数字格式化
        for r in range(5, end_row + 1):
            ws.cell(row=r, column=4).number_format = '#,##0.00'
            ws.cell(row=r, column=5).number_format = '#,##0'
            ws.cell(row=r, column=6).number_format = '0.0'
            ws.cell(row=r, column=7).number_format = '#,##0'

        _auto_width(ws, len(headers))

        # 冻结表头
        ws.freeze_panes = "A5"

        return self._save(wb)

    # ========== 2. 导出分析报告 ==========

    def export_analysis_report(self) -> io.BytesIO:
        """导出评分 + 类目分析报告"""
        wb = Workbook()
        products = product_service.get_all_raw_products()

        # ---- Sheet 1: 评分排名 ----
        ws1 = wb.active
        ws1.title = "智能评分排名"

        from app.services.scoring_service import scoring_service
        scores = scoring_service.score_all_products(products)

        ws1.merge_cells("A1:H1")
        ws1.cell(row=1, column=1, value="智能选品评分排名报告").font = TITLE_FONT

        headers1 = ["排名", "商品ID", "商品标题", "综合得分", "等级", "销量得分", "评分得分", "评论得分", "价格得分"]
        for col, h in enumerate(headers1, 1):
            ws1.cell(row=3, column=col, value=h)
        _style_header(ws1, 3, len(headers1))

        for i, s in enumerate(scores):
            row = 4 + i
            product = next((p for p in products if p["id"] == s.product_id), None)
            ws1.cell(row=row, column=1, value=i + 1)
            ws1.cell(row=row, column=2, value=s.product_id)
            ws1.cell(row=row, column=3, value=product["title"][:60] if product else "")
            ws1.cell(row=row, column=4, value=s.composite_score)
            ws1.cell(row=row, column=5, value=s.tier)
            ws1.cell(row=row, column=6, value=s.dimensions.sales_score)
            ws1.cell(row=row, column=7, value=s.dimensions.rating_score)
            ws1.cell(row=row, column=8, value=s.dimensions.reviews_score)
            ws1.cell(row=row, column=9, value=s.dimensions.price_score)

        _style_data(ws1, 4, 3 + len(scores), len(headers1))
        _auto_width(ws1, len(headers1))
        ws1.freeze_panes = "A4"

        # ---- Sheet 2: 类目汇总 ----
        ws2 = wb.create_sheet("类目汇总")

        by_cat: dict[str, list] = {}
        for p in products:
            by_cat.setdefault(p["category"], []).append(p)

        ws2.merge_cells("A1:F1")
        ws2.cell(row=1, column=1, value="类目汇总分析").font = TITLE_FONT

        headers2 = ["类目", "商品数量", "平均价格($)", "平均评分", "总销量", "预估收入($)"]
        for col, h in enumerate(headers2, 1):
            ws2.cell(row=3, column=col, value=h)
        _style_header(ws2, 3, len(headers2))

        row = 4
        for cat, prods in sorted(by_cat.items()):
            ws2.cell(row=row, column=1, value=cat)
            ws2.cell(row=row, column=2, value=len(prods))
            ws2.cell(row=row, column=3, value=round(sum(p["price"] for p in prods) / len(prods), 2))
            ws2.cell(row=row, column=4, value=round(sum(p["rating"] for p in prods) / len(prods), 2))
            ws2.cell(row=row, column=5, value=sum(p["sales"] for p in prods))
            ws2.cell(row=row, column=6, value=round(sum(p["price"] * p["sales"] for p in prods), 2))
            row += 1

        _style_data(ws2, 4, row - 1, len(headers2))
        _auto_width(ws2, len(headers2))
        ws2.freeze_panes = "A4"

        return self._save(wb)

    # ========== 3. 导出利润报表 ==========

    def export_profit_report(
        self,
        product_cost: float = 15,
        shipping: float = 5,
        commission_rate: float = 0.15,
        advertising_cost: float = 3,
    ) -> io.BytesIO:
        """导出利润预测报表（含多场景分析）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "利润预测报表"

        from app.services.profit_service import ProfitInput, profit_calculator

        # 标题
        ws.merge_cells("A1:H1")
        ws.cell(row=1, column=1, value="利润预测分析报表").font = TITLE_FONT

        ws.merge_cells("A2:H2")
        ws.cell(row=2, column=1, value=f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}").font = SUBTITLE_FONT

        # ---- 基础参数 ----
        ws.merge_cells("A4:B4")
        ws.cell(row=4, column=1, value="基础参数").font = Font(name="微软雅黑", size=12, bold=True)
        params = [
            ("商品成本($)", product_cost),
            ("运费($)", shipping),
            ("平台佣金率", f"{commission_rate * 100}%"),
            ("广告费用($)", advertising_cost),
        ]
        for i, (label, val) in enumerate(params):
            ws.cell(row=5 + i, column=1, value=label).font = DATA_FONT
            ws.cell(row=5 + i, column=2, value=val).font = DATA_FONT

        # ---- 多场景利润计算 ----
        ws.merge_cells("A10:H10")
        ws.cell(row=10, column=1, value="多场景价格分析").font = Font(name="微软雅黑", size=12, bold=True)

        scenarios = [
            ("保守定价", product_cost * 2.0),
            ("市场均价", product_cost * 2.5),
            ("进取定价", product_cost * 3.0),
            ("高端定价", product_cost * 4.0),
        ]

        headers = ["场景", "售价($)", "佣金($)", "总成本($)", "毛利润($)", "利润率(%)", "ROI(%)", "等级"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=12, column=col, value=h)
        _style_header(ws, 12, len(headers))

        for i, (scenario, price) in enumerate(scenarios):
            result = profit_calculator.calculate(ProfitInput(
                product_cost=product_cost, shipping=shipping,
                commission_rate=commission_rate, advertising_cost=advertising_cost,
                selling_price=price,
            ))
            row = 13 + i
            ws.cell(row=row, column=1, value=scenario)
            ws.cell(row=row, column=2, value=round(price, 2))
            ws.cell(row=row, column=3, value=result.commission_amount)
            ws.cell(row=row, column=4, value=result.total_cost)
            ws.cell(row=row, column=5, value=result.gross_profit)
            ws.cell(row=row, column=6, value=result.profit_margin)
            ws.cell(row=row, column=7, value=result.roi)
            ws.cell(row=row, column=8, value=result.profit_level)

        _style_data(ws, 13, 16, len(headers))
        _auto_width(ws, len(headers))
        ws.freeze_panes = "A13"

        return self._save(wb)

    # ========== 工具方法 ==========

    @staticmethod
    def _save(wb: Workbook) -> io.BytesIO:
        """将工作簿保存到内存流"""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# 全局导出服务单例
export_service = ExportService()
